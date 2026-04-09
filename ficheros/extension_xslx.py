"""
Archivos con extensión xlsx
Para leer archivos Excel necesitamos instalar la librería xlrd (u otras alternativas). Ejemplo:
"""

# from openpyxl import Workbook, load_workbook
import openpyxl

"""
1. Gestión del Libro (Workbook)
Para empezar, necesitas o crear un archivo nuevo o cargar uno que ya exista.

    Workbook(): Crea un libro nuevo en memoria.
    load_workbook('archivo.xlsx'): Abre un archivo existente.
    save('nombre.xlsx'): Guarda los cambios. Importante: Si no guardas, todo lo que hagas en el código se perderá al cerrar el script.
"""

""" 
# 1. Crear el libro y la hoja
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Empleados"

# 2. Crear el encabezado con estilo
encabezados = ["ID", "Nombre", "Departamento", "Salario", "Fecha de Ingreso"]
ws.append(encabezados)
 """

"""
2. Manejo de Hojas (Sheets)
Un libro puede tener múltiples hojas. Tienes que decirle a Python en cuál quieres escribir.

    wb.active: Selecciona la hoja que aparece abierta por defecto al abrir el Excel.
    wb.create_sheet("Nombre"): Crea una pestaña nueva.
    wb["Nombre"]: Accede a una hoja específica por su nombre.
    wb.sheetnames: Devuelve una lista con los nombres de todas las hojas.
"""

"""
3. Manipulación de Celdas y Datos
Aquí es donde ocurre la magia. Hay dos formas principales de interactuar con las celdas:

Acceso Directo
    sheet['A1'] = "Valor": Usa la nomenclatura estándar de Excel.
    sheet.cell(row=1, column=1, value="Valor"): Es mucho más útil cuando trabajas con bucles for, ya que usas números para filas y columnas.

Añadir datos masivos
    sheet.append([dato1, dato2, dato3]): Este es uno de los mejores métodos. Añade una fila completa al final de los datos existentes. Es ideal para volcar listas de información de forma rápida.
"""

# 1. Cargar el archivo
wb = openpyxl.load_workbook("./files/MOCK_DATA.xlsx")

# 2. Obtener el nombre de la hoja activa
# (o puedes usar wb.sheetnames[0] para la primera pestaña)
hoja = wb.active
titulo_hoja = str(hoja.title)

cabeceras = [
    celda.value for celda in hoja[1]
]  # obtenemos las cabeceras, con listas comprimidas
print(cabeceras)
print(f"Nombre de la hoja: {titulo_hoja}")
print(f"Cabeceras: {cabeceras}")


# TAREA 1: Cambiar para el primer registro los datos de: 'movie_title', 'genre', 'release_date'
print(
    "\n-- # TAREA 1: Cambiar para el primer registro los datos de: 'movie_title', 'genre', 'release_date' --\n"
)
hoja.cell(row=2, column=1, value="Avatar")
hoja.cell(row=2, column=2, value="Ciencia Ficción")
hoja.cell(row=2, column=3, value="2024-05-20")

hoja["A3"] = "Doom"
hoja["B3"] = "Aventura"
hoja["C3"] = "2024-05-20"

hoja.append(["Spiderman", "Aventura", "2024-05-20"])  # agrega una fila al final

hoja.insert_rows(2)  # inserta una fila en la segunda posición
datos_nuevaFila = ["Superman", "Ciencia Ficción", "2024-05-20"]

for i, valor in enumerate(datos_nuevaFila, start=1):
    hoja.cell(row=2, column=i, value=valor)


# 3. Guardar los cambios
# Te recomiendo guardarlo con otro nombre para no machacar el original mientras pruebas
wb.save("./files/MOCK_DATA_MODIFICADO.xlsx")

print("Primer registro actualizado correctamente.")

# 4. Cerrar el libro
wb.close()
